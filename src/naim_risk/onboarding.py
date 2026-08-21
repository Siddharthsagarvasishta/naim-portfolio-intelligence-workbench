"""Governed local data onboarding for canonical nAIM contracts.

The module intentionally keeps ingestion separate from active analytical data.  It
accepts only bounded, explicitly supported sources, evaluates transformations with
an AST interpreter (never ``eval``), quarantines invalid rows and requires an
approval transition before an import profile becomes active.
"""

from __future__ import annotations

import ast
import csv
import hashlib
import json
import math
import os
import re
import shutil
import sqlite3
import tempfile
import unicodedata
import uuid
import warnings
import zipfile
from collections.abc import Mapping, Sequence
from dataclasses import asdict, dataclass
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any, Literal

import duckdb
import numpy as np
import pandas as pd
import pyarrow as pa
import pyarrow.parquet as pq
from openpyxl import load_workbook
from sqlalchemy import create_engine, inspect, text
from sqlalchemy.engine import make_url

from naim_risk.onboarding_errors import (
    FormulaSafetyError,
    OnboardingError,
    ProfileApprovalError,
    SourceReadError,
    SourceSafetyError,
)
from naim_risk.workflow import (
    ConcurrencyConflict,
    DuplicateObject,
    ObjectNotFound,
    WorkflowStore,
)

SUPPORTED_FILE_SUFFIXES = {
    ".csv": "csv",
    ".xlsx": "xlsx",
    ".parquet": "parquet",
    ".json": "json",
    ".sqlite": "sqlite",
    ".sqlite3": "sqlite",
    ".db": "sqlite",
    ".duckdb": "duckdb",
}
DATABASE_KINDS = {"sqlite", "duckdb", "postgresql"}
PROFILE_OBJECT_TYPE = "configuration_change"
PROFILE_EXTERNAL_PREFIX = "onboarding-profile:"
PROFILE_ID_PATTERN = re.compile(r"^[A-Za-z0-9][A-Za-z0-9_.-]{0,63}$")
IDENTIFIER_PATTERN = re.compile(r"^[A-Za-z_][A-Za-z0-9_]{0,62}$")
ENVIRONMENT_PATTERN = re.compile(r"^NAIM_ONBOARDING_[A-Z0-9_]{1,100}_URL$")


@dataclass(frozen=True)
class ContractField:
    """One canonical field and its coercion rules."""

    name: str
    data_type: Literal["string", "integer", "number", "date", "datetime", "boolean"]
    required: bool = False
    non_negative: bool = False
    allowed_values: tuple[str, ...] = ()
    description: str = ""


@dataclass(frozen=True)
class DataContract:
    """Canonical onboarding contract used for mapping and validation."""

    contract_id: str
    version: str
    description: str
    fields: tuple[ContractField, ...]
    unique_key: tuple[str, ...]

    @property
    def field_lookup(self) -> dict[str, ContractField]:
        return {field.name: field for field in self.fields}


def _field(
    name: str,
    data_type: Literal["string", "integer", "number", "date", "datetime", "boolean"],
    *,
    required: bool = False,
    non_negative: bool = False,
    allowed_values: Sequence[str] = (),
    description: str = "",
) -> ContractField:
    return ContractField(
        name=name,
        data_type=data_type,
        required=required,
        non_negative=non_negative,
        allowed_values=tuple(allowed_values),
        description=description,
    )


CONTRACTS: dict[str, DataContract] = {
    "account_master": DataContract(
        "account_master",
        "1.0.0",
        "One point-in-time row per governed account.",
        (
            _field("account_id", "string", required=True),
            _field("opened_date", "date"),
            _field("closed_date", "date"),
            _field("product", "string"),
            _field("region", "string"),
            _field("acquisition_channel", "string"),
            _field("partner_id", "string"),
            _field("vendor_id", "string"),
            _field("membership_tier", "string"),
            _field("credit_limit", "number", non_negative=True),
            _field("active", "boolean"),
        ),
        ("account_id",),
    ),
    "account_month_performance": DataContract(
        "account_month_performance",
        "1.0.0",
        "Monthly account performance with exposure and loss measures.",
        (
            _field("account_id", "string", required=True),
            _field("month", "date", required=True),
            _field("exposure", "number", non_negative=True),
            _field("balance", "number", non_negative=True),
            _field("net_loss", "number"),
            _field("fraud_loss", "number", non_negative=True),
            _field("revenue", "number"),
            _field("delinquency_rate", "number", non_negative=True),
            _field("utilisation", "number", non_negative=True),
        ),
        ("account_id", "month"),
    ),
    "strategy_decision": DataContract(
        "strategy_decision",
        "1.0.0",
        "Governed strategy decisions at their decision timestamp.",
        (
            _field("decision_id", "string", required=True),
            _field("account_id", "string", required=True),
            _field("decision_date", "datetime", required=True),
            _field("strategy", "string", required=True),
            _field("action", "string", required=True),
            _field("decision_score", "number"),
            _field("approved", "boolean"),
            _field("model_version", "string"),
        ),
        ("decision_id",),
    ),
    "partner_performance": DataContract(
        "partner_performance",
        "1.0.0",
        "Periodic partner portfolio performance.",
        (
            _field("partner_id", "string", required=True),
            _field("period", "date", required=True),
            _field("accounts", "integer", non_negative=True),
            _field("exposure", "number", non_negative=True),
            _field("net_loss", "number"),
            _field("revenue", "number"),
            _field("complaints", "integer", non_negative=True),
        ),
        ("partner_id", "period"),
    ),
    "vendor_performance": DataContract(
        "vendor_performance",
        "1.0.0",
        "Periodic servicing or collections vendor performance.",
        (
            _field("vendor_id", "string", required=True),
            _field("period", "date", required=True),
            _field("volume", "integer", non_negative=True),
            _field("recoveries", "number", non_negative=True),
            _field("cost", "number", non_negative=True),
            _field("service_level", "number", non_negative=True),
            _field("complaints", "integer", non_negative=True),
        ),
        ("vendor_id", "period"),
    ),
    "membership_history": DataContract(
        "membership_history",
        "1.0.0",
        "Effective-dated account membership history.",
        (
            _field("membership_id", "string", required=True),
            _field("account_id", "string", required=True),
            _field("start_date", "date", required=True),
            _field("end_date", "date"),
            _field("membership_tier", "string", required=True),
            _field("status", "string"),
            _field("fee", "number", non_negative=True),
        ),
        ("membership_id",),
    ),
    "benefit_usage": DataContract(
        "benefit_usage",
        "1.0.0",
        "Account-level use of funded membership benefits.",
        (
            _field("usage_id", "string", required=True),
            _field("account_id", "string", required=True),
            _field("usage_date", "datetime", required=True),
            _field("benefit", "string", required=True),
            _field("units", "number", non_negative=True),
            _field("funded_amount", "number", non_negative=True),
            _field("vendor_id", "string"),
        ),
        ("usage_id",),
    ),
    "economic_assumptions": DataContract(
        "economic_assumptions",
        "1.0.0",
        "Versioned macroeconomic scenario assumptions.",
        (
            _field("scenario", "string", required=True),
            _field("period", "date", required=True),
            _field("assumption_version", "string", required=True),
            _field("unemployment_rate", "number", non_negative=True),
            _field("base_rate", "number"),
            _field("gdp_growth", "number"),
            _field("inflation_rate", "number"),
            _field("house_price_growth", "number"),
        ),
        ("scenario", "period", "assumption_version"),
    ),
}


def list_contracts() -> list[dict[str, Any]]:
    """Return portable metadata for all eight canonical contracts."""

    return [
        {
            "contract_id": contract.contract_id,
            "version": contract.version,
            "description": contract.description,
            "unique_key": list(contract.unique_key),
            "fields": [asdict(field) for field in contract.fields],
        }
        for contract in CONTRACTS.values()
    ]


def _is_null(value: Any) -> bool:
    if value is None:
        return True
    try:
        result = pd.isna(value)
    except (TypeError, ValueError):
        return False
    return bool(result) if isinstance(result, (bool, np.bool_)) else False


def _safe_number(value: Any) -> int | float | None:
    if _is_null(value):
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, np.integer)):
        return int(value)
    if isinstance(value, (float, np.floating)):
        converted = float(value)
        return converted if math.isfinite(converted) else None
    try:
        converted = float(value)
    except (OverflowError, TypeError, ValueError):
        return None
    return converted if math.isfinite(converted) else None


def _date_diff(unit: Any, start: Any, end: Any) -> int | float | None:
    if _is_null(start) or _is_null(end):
        return None
    unit_name = str(unit).strip().lower()
    if unit_name not in {"days", "months", "years"}:
        raise FormulaSafetyError("date_diff unit must be days, months or years")
    try:
        start_value = pd.Timestamp(start)
        end_value = pd.Timestamp(end)
    except (TypeError, ValueError) as exc:
        raise FormulaSafetyError("date_diff received an invalid date") from exc
    if unit_name == "days":
        return (end_value - start_value).total_seconds() / 86_400
    months = (end_value.year - start_value.year) * 12 + end_value.month - start_value.month
    if unit_name == "months":
        return months
    return months / 12


def _normalise_string(value: Any) -> str | None:
    if _is_null(value):
        return None
    normalised = unicodedata.normalize("NFKC", str(value))
    return " ".join(normalised.strip().split()).casefold()


def _truthy(value: Any) -> bool:
    return False if _is_null(value) else bool(value)


class SafeFormula:
    """Validated, non-evaluating AST interpreter for one derived-field formula."""

    _binary_operators = {
        ast.Add: lambda left, right: left + right,
        ast.Sub: lambda left, right: left - right,
        ast.Mult: lambda left, right: left * right,
        ast.Mod: lambda left, right: left % right,
        ast.Pow: lambda left, right: left**right,
    }
    _comparison_operators = {
        ast.Eq: lambda left, right: left == right,
        ast.NotEq: lambda left, right: left != right,
        ast.Lt: lambda left, right: left < right,
        ast.LtE: lambda left, right: left <= right,
        ast.Gt: lambda left, right: left > right,
        ast.GtE: lambda left, right: left >= right,
        ast.In: lambda left, right: left in right,
        ast.NotIn: lambda left, right: left not in right,
    }
    _allowed_functions = {
        "category_map",
        "clip",
        "coalesce",
        "date_diff",
        "if_else",
        "map_value",
        "normalize",
    }
    _allowed_nodes = (
        ast.Expression,
        ast.BinOp,
        ast.UnaryOp,
        ast.BoolOp,
        ast.Compare,
        ast.Call,
        ast.Name,
        ast.Load,
        ast.Constant,
        ast.Dict,
        ast.List,
        ast.Tuple,
        ast.Add,
        ast.Sub,
        ast.Mult,
        ast.Div,
        ast.Mod,
        ast.Pow,
        ast.USub,
        ast.UAdd,
        ast.Not,
        ast.And,
        ast.Or,
        ast.Eq,
        ast.NotEq,
        ast.Lt,
        ast.LtE,
        ast.Gt,
        ast.GtE,
        ast.In,
        ast.NotIn,
    )

    def __init__(
        self,
        expression: str,
        *,
        allowed_fields: Sequence[str],
        max_nodes: int = 120,
    ) -> None:
        if not isinstance(expression, str) or not expression.strip():
            raise FormulaSafetyError("A non-empty formula is required")
        if len(expression) > 1_000:
            raise FormulaSafetyError("Formula exceeds the 1,000 character limit")
        try:
            tree = ast.parse(expression, mode="eval")
        except SyntaxError as exc:
            raise FormulaSafetyError("Formula syntax is invalid") from exc
        nodes = list(ast.walk(tree))
        if len(nodes) > max_nodes:
            raise FormulaSafetyError("Formula is too complex")
        allowed_field_set = set(allowed_fields)
        for node in nodes:
            if not isinstance(node, self._allowed_nodes):
                raise FormulaSafetyError(
                    f"Formula construct {type(node).__name__} is not permitted"
                )
            if isinstance(node, ast.Name):
                if node.id.startswith("_"):
                    raise FormulaSafetyError("Private or special names are not permitted")
                if node.id not in allowed_field_set and node.id not in self._allowed_functions:
                    raise FormulaSafetyError(f"Unknown formula field or function: {node.id}")
            if isinstance(node, ast.Call):
                if (
                    not isinstance(node.func, ast.Name)
                    or node.func.id not in self._allowed_functions
                ):
                    raise FormulaSafetyError("Only approved formula functions may be called")
                if node.keywords:
                    raise FormulaSafetyError("Formula function keyword arguments are not permitted")
            if isinstance(node, (ast.Dict, ast.List, ast.Tuple)):
                if len(getattr(node, "elts", getattr(node, "keys", []))) > 100:
                    raise FormulaSafetyError(
                        "Formula collection literals may contain at most 100 items"
                    )
        self.expression = expression
        self._tree = tree

    @staticmethod
    def _call(name: str, arguments: list[Any]) -> Any:
        if name == "coalesce":
            if not arguments:
                raise FormulaSafetyError("coalesce requires at least one argument")
            return next((value for value in arguments if not _is_null(value)), None)
        if name == "if_else":
            if len(arguments) != 3:
                raise FormulaSafetyError("if_else requires condition, true value and false value")
            return arguments[1] if _truthy(arguments[0]) else arguments[2]
        if name in {"map_value", "category_map"}:
            if len(arguments) not in {2, 3} or not isinstance(arguments[1], dict):
                raise FormulaSafetyError("map_value requires a value, mapping and optional default")
            default = arguments[2] if len(arguments) == 3 else None
            return arguments[1].get(arguments[0], default)
        if name == "clip":
            if len(arguments) != 3:
                raise FormulaSafetyError("clip requires a value, lower bound and upper bound")
            if _is_null(arguments[0]):
                return None
            value, lower, upper = (_safe_number(item) for item in arguments)
            if value is None or lower is None or upper is None:
                return None
            if lower > upper:
                raise FormulaSafetyError("clip lower bound cannot exceed upper bound")
            return min(max(value, lower), upper)
        if name == "normalize":
            if len(arguments) != 1:
                raise FormulaSafetyError("normalize requires one argument")
            return _normalise_string(arguments[0])
        if name == "date_diff":
            if len(arguments) != 3:
                raise FormulaSafetyError("date_diff requires unit, start and end")
            return _date_diff(*arguments)
        raise FormulaSafetyError(f"Unsupported formula function: {name}")

    def _interpret(self, node: ast.AST, row: Mapping[str, Any]) -> Any:
        if isinstance(node, ast.Expression):
            return self._interpret(node.body, row)
        if isinstance(node, ast.Constant):
            if not isinstance(node.value, (str, int, float, bool, type(None))):
                raise FormulaSafetyError("Unsupported formula literal")
            return node.value
        if isinstance(node, ast.Name):
            if node.id not in row:
                raise FormulaSafetyError(f"Formula field is absent from this source: {node.id}")
            return row[node.id]
        if isinstance(node, ast.List):
            return [self._interpret(item, row) for item in node.elts]
        if isinstance(node, ast.Tuple):
            return tuple(self._interpret(item, row) for item in node.elts)
        if isinstance(node, ast.Dict):
            return {
                self._interpret(key, row): self._interpret(value, row)
                for key, value in zip(node.keys, node.values, strict=True)
            }
        if isinstance(node, ast.UnaryOp):
            value = self._interpret(node.operand, row)
            if isinstance(node.op, ast.Not):
                return not _truthy(value)
            if _is_null(value):
                return None
            if isinstance(node.op, ast.USub):
                return -value
            if isinstance(node.op, ast.UAdd):
                return +value
        if isinstance(node, ast.BoolOp):
            values = [_truthy(self._interpret(value, row)) for value in node.values]
            return all(values) if isinstance(node.op, ast.And) else any(values)
        if isinstance(node, ast.Compare):
            left = self._interpret(node.left, row)
            for operator, comparator in zip(node.ops, node.comparators, strict=True):
                right = self._interpret(comparator, row)
                operation = self._comparison_operators.get(type(operator))
                if operation is None:
                    raise FormulaSafetyError("Unsupported comparison operator")
                if _is_null(left) or _is_null(right):
                    both_null = _is_null(left) and _is_null(right)
                    if isinstance(operator, ast.Eq):
                        comparison_result = both_null
                    elif isinstance(operator, ast.NotEq):
                        comparison_result = not both_null
                    else:
                        comparison_result = False
                    if not comparison_result:
                        return False
                    left = right
                    continue
                try:
                    if not operation(left, right):
                        return False
                except (TypeError, ValueError):
                    return False
                left = right
            return True
        if isinstance(node, ast.BinOp):
            left = self._interpret(node.left, row)
            right = self._interpret(node.right, row)
            if _is_null(left) or _is_null(right):
                return None
            left_number = _safe_number(left)
            right_number = _safe_number(right)
            if left_number is None or right_number is None:
                return None
            if isinstance(node.op, ast.Div):
                return None if right_number == 0 else left_number / right_number
            if isinstance(node.op, ast.Mod) and right_number == 0:
                return None
            if isinstance(node.op, ast.Pow) and (
                abs(right_number) > 12 or abs(left_number) > 1e100
            ):
                raise FormulaSafetyError("Exponentiation exceeds the formula safety bound")
            operation = self._binary_operators.get(type(node.op))
            if operation is None:
                raise FormulaSafetyError("Unsupported arithmetic operator")
            try:
                result = operation(left_number, right_number)
            except (ArithmeticError, TypeError, ValueError, OverflowError):
                return None
            if isinstance(result, (int, float)) and not math.isfinite(float(result)):
                return None
            return result
        if isinstance(node, ast.Call) and isinstance(node.func, ast.Name):
            arguments = [self._interpret(argument, row) for argument in node.args]
            return self._call(node.func.id, arguments)
        raise FormulaSafetyError(f"Formula construct {type(node).__name__} is not permitted")

    def evaluate(self, row: Mapping[str, Any]) -> Any:
        """Interpret the validated formula against one source row."""

        try:
            return self._interpret(self._tree, row)
        except FormulaSafetyError:
            raise
        except Exception as exc:
            raise FormulaSafetyError("Formula could not be evaluated safely") from exc


def _portable_value(value: Any) -> Any:
    """Convert pandas/numpy scalars into deterministic JSON values."""

    if _is_null(value):
        return None
    if isinstance(value, (pd.Timestamp, datetime)):
        stamp = pd.Timestamp(value)
        if stamp.tzinfo is None:
            stamp = stamp.tz_localize(UTC)
        return stamp.tz_convert(UTC).isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (np.integer,)):
        return int(value)
    if isinstance(value, (np.floating,)):
        converted = float(value)
        return converted if math.isfinite(converted) else None
    if isinstance(value, (np.bool_,)):
        return bool(value)
    if isinstance(value, bytes):
        return value.hex()
    return value


def _portable_records(frame: pd.DataFrame, limit: int | None = None) -> list[dict[str, Any]]:
    selected = frame if limit is None else frame.head(limit)
    return [
        {str(key): _portable_value(value) for key, value in row.items()}
        for row in selected.to_dict(orient="records")
    ]


def _infer_series_type(series: pd.Series) -> tuple[str, float]:
    non_null = series.dropna()
    if non_null.empty:
        return "null", 1.0
    if pd.api.types.is_bool_dtype(non_null):
        return "boolean", 1.0
    if pd.api.types.is_integer_dtype(non_null):
        return "integer", 1.0
    if pd.api.types.is_numeric_dtype(non_null):
        return "number", 1.0
    strings = non_null.astype(str).str.strip()
    lowered = strings.str.casefold()
    boolean_ratio = lowered.isin({"true", "false", "yes", "no", "y", "n", "0", "1"}).mean()
    numeric = pd.to_numeric(strings, errors="coerce")
    numeric_ratio = numeric.notna().mean()
    integer_ratio = (
        (numeric.dropna() % 1 == 0).mean() * numeric_ratio if numeric.notna().any() else 0.0
    )
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        dates = pd.to_datetime(strings, errors="coerce", utc=True)
    date_ratio = dates.notna().mean()
    if boolean_ratio >= 0.98:
        return "boolean", float(boolean_ratio)
    if numeric_ratio >= 0.98:
        inferred = "integer" if integer_ratio >= 0.98 else "number"
        return inferred, float(integer_ratio if inferred == "integer" else numeric_ratio)
    if date_ratio >= 0.98:
        return "datetime", float(date_ratio)
    return "string", 1.0


def _safe_identifier(value: str, *, label: str = "identifier") -> str:
    if not IDENTIFIER_PATTERN.fullmatch(value):
        raise SourceSafetyError(
            f"{label} must begin with a letter or underscore and contain only letters, numbers "
            "and underscores"
        )
    return value


def _quote_identifier_path(value: str) -> str:
    parts = value.split(".")
    if len(parts) > 2 or not parts:
        raise SourceSafetyError("Database table must be table or schema.table")
    return ".".join(f'"{_safe_identifier(part, label="database table")}"' for part in parts)


def _atomic_json(path: Path, payload: Mapping[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(
        "w",
        encoding="utf-8",
        dir=path.parent,
        prefix=f".{path.name}.",
        suffix=".tmp",
        delete=False,
    ) as handle:
        json.dump(payload, handle, indent=2, sort_keys=True, default=_portable_value)
        handle.write("\n")
        temporary = Path(handle.name)
    os.replace(temporary, path)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _profile_identifier(value: str) -> str:
    if not PROFILE_ID_PATTERN.fullmatch(value):
        raise ValueError(
            "profile_id must be 1-64 characters using only letters, numbers, '.', '_' or '-'"
        )
    return value


def _source_kind_from_path(path: Path) -> str:
    kind = SUPPORTED_FILE_SUFFIXES.get(path.suffix.casefold())
    if kind is None:
        supported = ", ".join(sorted(SUPPORTED_FILE_SUFFIXES))
        raise SourceSafetyError(f"Unsupported source suffix; expected one of: {supported}")
    return kind


def _boolean_series(series: pd.Series) -> tuple[pd.Series, pd.Series]:
    mapping: dict[Any, bool] = {
        True: True,
        False: False,
        "true": True,
        "false": False,
        "yes": True,
        "no": False,
        "y": True,
        "n": False,
        "1": True,
        "0": False,
    }
    converted = series.map(
        lambda value: (
            None
            if _is_null(value)
            else mapping.get(str(value).strip().casefold() if isinstance(value, str) else value)
        )
    )
    invalid = series.notna() & converted.isna()
    return converted.astype("boolean"), invalid


@dataclass(frozen=True)
class _ValidationFrames:
    canonical: pd.DataFrame
    valid: pd.DataFrame
    invalid: pd.DataFrame
    source_row_numbers: pd.Series
    errors: tuple[dict[str, Any], ...]
    error_rows: frozenset[int]
    summary: dict[str, Any]


class OnboardingStudio:
    """Facade for bounded source inspection, validation, loading and approval."""

    def __init__(
        self,
        onboarding_root: str | Path,
        *,
        workflow_store: WorkflowStore | None = None,
        max_upload_bytes: int = 50 * 1024 * 1024,
        max_rows: int = 1_000_000,
        max_columns: int = 250,
        max_preview_rows: int = 200,
        max_error_preview: int = 200,
    ) -> None:
        root = Path(onboarding_root).expanduser()
        root.mkdir(parents=True, exist_ok=True)
        self.root = root.resolve()
        self.sources_root = self.root / "sources"
        self.profiles_root = self.root / "profiles"
        self.namespace_root = self.root / "namespace"
        self.quarantine_root = self.root / "quarantine"
        self.preview_marts_root = self.root / "preview_marts"
        for directory in (
            self.sources_root,
            self.profiles_root,
            self.namespace_root,
            self.quarantine_root,
            self.preview_marts_root,
        ):
            directory.mkdir(parents=True, exist_ok=True, mode=0o750)
        if max_upload_bytes < 1 or max_rows < 1 or max_columns < 1:
            raise ValueError("Onboarding safety limits must be positive")
        self.workflow_store = workflow_store
        self.max_upload_bytes = int(max_upload_bytes)
        self.max_rows = int(max_rows)
        self.max_columns = int(max_columns)
        self.max_preview_rows = max(1, min(int(max_preview_rows), 2_000))
        self.max_error_preview = max(1, min(int(max_error_preview), 5_000))

    @staticmethod
    def contracts() -> list[dict[str, Any]]:
        """Return the eight contract definitions."""

        return list_contracts()

    def _contained(self, base: Path, candidate: Path) -> Path:
        resolved = candidate.resolve()
        try:
            resolved.relative_to(base.resolve())
        except ValueError as exc:
            raise SourceSafetyError("Source path leaves the governed onboarding root") from exc
        return resolved

    @staticmethod
    def _reject_symlink_components(base: Path, candidate: Path) -> None:
        try:
            relative = candidate.absolute().relative_to(base.absolute())
        except ValueError as exc:
            raise SourceSafetyError("Source path leaves the governed onboarding root") from exc
        current = base.absolute()
        for part in relative.parts:
            current = current / part
            if current.is_symlink():
                raise SourceSafetyError("Symbolic-link sources are not permitted")

    def _source_path(self, source: Mapping[str, Any]) -> Path:
        relative = source.get("relative_path")
        if not isinstance(relative, str) or not relative:
            raise SourceSafetyError("A local source requires relative_path")
        if Path(relative).is_absolute() or ".." in Path(relative).parts:
            raise SourceSafetyError("Source paths must be relative and cannot contain '..'")
        candidate = self.root / relative
        self._reject_symlink_components(self.root, candidate)
        path = self._contained(self.root, candidate)
        self._contained(self.sources_root, path)
        if not path.is_file():
            raise SourceReadError("Selected onboarding source does not exist")
        expected_hash = source.get("sha256")
        if not isinstance(expected_hash, str) or not re.fullmatch(r"[0-9a-f]{64}", expected_hash):
            raise SourceSafetyError("Local source requires a registered SHA-256 hash")
        if _sha256(path) != expected_hash:
            raise SourceSafetyError("Source content no longer matches its registered hash")
        return path

    def _source_payload(
        self,
        path: Path,
        *,
        source_id: str | None = None,
        table: str | None = None,
        sheet: str | int | None = None,
    ) -> dict[str, Any]:
        source_id = source_id or path.parent.name
        kind = _source_kind_from_path(path)
        payload: dict[str, Any] = {
            "source_id": source_id,
            "kind": kind,
            "display_name": path.name,
            "relative_path": path.relative_to(self.root).as_posix(),
            "size_bytes": path.stat().st_size,
            "sha256": _sha256(path),
        }
        if table is not None:
            _quote_identifier_path(table)
            payload["table"] = table
        if sheet is not None:
            payload["sheet"] = sheet
        return payload

    def upload_source(self, filename: str, content: bytes) -> dict[str, Any]:
        """Store a bounded upload under a generated, non-user-controlled directory."""

        if not isinstance(content, bytes):
            raise TypeError("Uploaded content must be bytes")
        if not filename or Path(filename).name != filename or Path(filename).is_absolute():
            raise SourceSafetyError("Uploaded filename must be a plain filename without a path")
        if "\x00" in filename:
            raise SourceSafetyError("Uploaded filename contains an invalid character")
        if len(content) == 0:
            raise SourceReadError("Uploaded source is empty")
        if len(content) > self.max_upload_bytes:
            raise SourceSafetyError(
                f"Uploaded source exceeds the {self.max_upload_bytes} byte safety limit"
            )
        _source_kind_from_path(Path(filename))
        source_id = uuid.uuid4().hex
        destination_directory = self.sources_root / source_id
        destination_directory.mkdir(mode=0o750)
        destination = destination_directory / filename
        try:
            with destination.open("xb") as handle:
                handle.write(content)
            destination.chmod(0o640)
            source = self._source_payload(destination, source_id=source_id)
            # Read a single row now so malformed or suffix-spoofed uploads are not registered.
            if source["kind"] not in DATABASE_KINDS:
                self._read_source(source, limit=1)
            else:
                self.list_database_tables(source)
            _atomic_json(destination_directory / "source.json", source)
            return source
        except Exception:
            shutil.rmtree(destination_directory, ignore_errors=True)
            raise

    def select_source(
        self,
        relative_path: str,
        *,
        table: str | None = None,
        sheet: str | int | None = None,
    ) -> dict[str, Any]:
        """Select an existing file only from the governed ``sources`` namespace."""

        path_value = Path(relative_path)
        if path_value.is_absolute() or ".." in path_value.parts:
            raise SourceSafetyError("Selected source path must be a safe relative path")
        candidate = self.sources_root / path_value
        self._reject_symlink_components(self.sources_root, candidate)
        path = self._contained(self.sources_root, candidate)
        if not path.is_file():
            raise SourceReadError("Selected onboarding source is not a regular file")
        if path.stat().st_size > self.max_upload_bytes:
            raise SourceSafetyError("Selected source exceeds the configured safety limit")
        return self._source_payload(path, table=table, sheet=sheet)

    def configure_postgresql_source(
        self,
        *,
        url_env: str,
        table: str,
        source_id: str | None = None,
    ) -> dict[str, Any]:
        """Return a secret-free PostgreSQL source descriptor.

        Connection URLs are accepted only through a named environment variable and
        are never copied into profile, response or audit payloads.
        """

        if not ENVIRONMENT_PATTERN.fullmatch(url_env):
            raise SourceSafetyError(
                "PostgreSQL URL references must use a NAIM_ONBOARDING_*_URL environment variable"
            )
        _quote_identifier_path(table)
        return {
            "source_id": source_id or uuid.uuid4().hex,
            "kind": "postgresql",
            "display_name": f"PostgreSQL:{table}",
            "url_env": url_env,
            "table": table,
        }

    def with_table(self, source: Mapping[str, Any], table: str) -> dict[str, Any]:
        """Bind a safe table name to a database source descriptor."""

        _quote_identifier_path(table)
        selected = dict(source)
        if selected.get("kind") not in DATABASE_KINDS:
            raise ValueError("Table selection applies only to database sources")
        selected["table"] = table
        return selected

    def list_database_tables(self, source: Mapping[str, Any]) -> list[str]:
        """List base tables without executing user-provided SQL."""

        kind = str(source.get("kind", ""))
        if kind == "sqlite":
            path = self._source_path(source)
            try:
                connection = sqlite3.connect(f"file:{path}?mode=ro", uri=True)
                try:
                    rows = connection.execute(
                        "SELECT name FROM sqlite_master "
                        "WHERE type = 'table' AND name NOT LIKE 'sqlite_%' ORDER BY name"
                    ).fetchall()
                finally:
                    connection.close()
            except sqlite3.DatabaseError as exc:
                raise SourceReadError("SQLite source is malformed or unreadable") from exc
            return [str(row[0]) for row in rows]
        if kind == "duckdb":
            path = self._source_path(source)
            try:
                connection = duckdb.connect(
                    str(path), read_only=True, config={"enable_external_access": "false"}
                )
                try:
                    rows = connection.execute(
                        "SELECT schema_name, table_name FROM duckdb_tables() "
                        "WHERE internal = false ORDER BY schema_name, table_name"
                    ).fetchall()
                finally:
                    connection.close()
            except duckdb.Error as exc:
                raise SourceReadError("DuckDB source is malformed or unreadable") from exc
            return [
                str(table) if str(schema) in {"main", "memory"} else f"{schema}.{table}"
                for schema, table in rows
            ]
        if kind == "postgresql":
            engine = self._postgresql_engine(source)
            try:
                inspector = inspect(engine)
                return sorted(
                    f"public.{table}" for table in inspector.get_table_names(schema="public")
                )
            except Exception as exc:
                raise SourceReadError(
                    "PostgreSQL source could not list tables using its environment reference"
                ) from exc
            finally:
                engine.dispose()
        raise ValueError("Table listing requires a SQLite, DuckDB or PostgreSQL source")

    @staticmethod
    def _postgresql_engine(source: Mapping[str, Any]):
        url_env = source.get("url_env")
        if not isinstance(url_env, str) or not ENVIRONMENT_PATTERN.fullmatch(url_env):
            raise SourceSafetyError(
                "PostgreSQL source requires a safe NAIM_ONBOARDING_*_URL reference"
            )
        raw_url = os.getenv(url_env)
        if not raw_url:
            raise SourceReadError(f"PostgreSQL environment reference {url_env} is not configured")
        try:
            parsed = make_url(raw_url)
        except Exception as exc:
            raise SourceReadError(
                "PostgreSQL environment reference contains an invalid URL"
            ) from exc
        if parsed.get_backend_name() not in {"postgresql", "postgres"}:
            raise SourceSafetyError("PostgreSQL source URL must use a PostgreSQL driver")
        try:
            return create_engine(raw_url, future=True, pool_pre_ping=True)
        except Exception as exc:
            raise SourceReadError("PostgreSQL connection configuration is invalid") from exc

    def _validate_frame(self, frame: pd.DataFrame) -> pd.DataFrame:
        if len(frame) > self.max_rows:
            raise SourceSafetyError(f"Source exceeds the {self.max_rows:,} row safety limit")
        if len(frame.columns) > self.max_columns:
            raise SourceSafetyError(f"Source exceeds the {self.max_columns:,} column safety limit")
        names = [str(column).strip() for column in frame.columns]
        if any(not name for name in names):
            raise SourceReadError("Source contains a blank column name")
        if len(names) != len(set(names)):
            raise SourceReadError("Source contains duplicate column names")
        if any(len(name) > 256 for name in names):
            raise SourceSafetyError("Source column names may contain at most 256 characters")
        cleaned = frame.copy()
        cleaned.columns = names
        return cleaned

    def _inspect_xlsx(self, path: Path, *, sheet: str | int | None) -> str | int:
        try:
            with zipfile.ZipFile(path) as archive:
                members = archive.infolist()
                expanded_size = sum(item.file_size for item in members)
                if expanded_size > max(self.max_upload_bytes * 20, 100 * 1024 * 1024):
                    raise SourceSafetyError("XLSX expanded content exceeds the safety limit")
                if any(
                    Path(item.filename).is_absolute() or ".." in Path(item.filename).parts
                    for item in members
                ):
                    raise SourceSafetyError("XLSX archive contains an unsafe internal path")
                if any(item.filename.startswith("xl/externalLinks/") for item in members):
                    raise SourceSafetyError("XLSX external links are not permitted")
                if any(
                    item.filename == "xl/vbaProject.bin"
                    or item.filename.startswith("xl/embeddings/")
                    or item.filename.startswith("xl/oleObjects/")
                    for item in members
                ):
                    raise SourceSafetyError("XLSX embedded executable objects are not permitted")
            workbook = load_workbook(path, read_only=True, data_only=False, keep_links=False)
        except SourceSafetyError:
            raise
        except (OSError, ValueError, zipfile.BadZipFile) as exc:
            raise SourceReadError("XLSX source is malformed or unreadable") from exc
        try:
            selected: str | int = 0 if sheet is None else sheet
            if isinstance(selected, int):
                if selected < 0 or selected >= len(workbook.sheetnames):
                    raise SourceReadError("Requested XLSX sheet index does not exist")
                worksheet = workbook[workbook.sheetnames[selected]]
            elif isinstance(selected, str) and selected in workbook.sheetnames:
                worksheet = workbook[selected]
            else:
                raise SourceReadError("Requested XLSX sheet does not exist")
            if worksheet.max_row > self.max_rows + 1:
                raise SourceSafetyError(f"XLSX exceeds the {self.max_rows:,} row safety limit")
            if worksheet.max_column > self.max_columns:
                raise SourceSafetyError(
                    f"XLSX exceeds the {self.max_columns:,} column safety limit"
                )
            header_values = [
                "" if cell.value is None else str(cell.value).strip()
                for cell in next(
                    worksheet.iter_rows(min_row=1, max_row=1, max_col=worksheet.max_column),
                    (),
                )
            ]
            if any(not value for value in header_values):
                raise SourceReadError("XLSX source contains a blank column name")
            if len(header_values) != len(set(header_values)):
                raise SourceReadError("XLSX source contains duplicate column names")
            for row in worksheet.iter_rows(
                min_row=1,
                max_row=min(worksheet.max_row, self.max_rows + 1),
                max_col=min(worksheet.max_column, self.max_columns),
            ):
                if any(cell.data_type == "f" for cell in row):
                    raise SourceSafetyError("XLSX formulas are not accepted as source data")
            return selected
        finally:
            workbook.close()

    def _read_local_tabular(
        self,
        source: Mapping[str, Any],
        *,
        limit: int,
    ) -> pd.DataFrame:
        path = self._source_path(source)
        kind = str(source.get("kind", ""))
        if _source_kind_from_path(path) != kind:
            raise SourceSafetyError("Source kind does not match its registered file suffix")
        try:
            if kind == "csv":
                with path.open("r", encoding="utf-8-sig", newline="") as handle:
                    try:
                        headers = [value.strip() for value in next(csv.reader(handle))]
                    except (StopIteration, csv.Error) as exc:
                        raise SourceReadError("CSV source has no readable header") from exc
                if any(not value for value in headers):
                    raise SourceReadError("CSV source contains a blank column name")
                if len(headers) != len(set(headers)):
                    raise SourceReadError("CSV source contains duplicate column names")
                frame = pd.read_csv(
                    path,
                    encoding="utf-8-sig",
                    nrows=limit,
                    on_bad_lines="error",
                    low_memory=False,
                    dtype=object,
                )
            elif kind == "xlsx":
                sheet = self._inspect_xlsx(path, sheet=source.get("sheet"))
                frame = pd.read_excel(path, sheet_name=sheet, nrows=limit, engine="openpyxl")
            elif kind == "parquet":
                parquet = pq.ParquetFile(path)
                if parquet.metadata.num_rows > self.max_rows:
                    raise SourceSafetyError(
                        f"Parquet source exceeds the {self.max_rows:,} row safety limit"
                    )
                if parquet.metadata.num_columns > self.max_columns:
                    raise SourceSafetyError(
                        f"Parquet source exceeds the {self.max_columns:,} column safety limit"
                    )
                batches = parquet.iter_batches(batch_size=min(limit, 65_536))
                tables: list[pa.Table] = []
                rows = 0
                for batch in batches:
                    remaining = limit - rows
                    if remaining <= 0:
                        break
                    table = pa.Table.from_batches([batch])
                    tables.append(table.slice(0, remaining))
                    rows += min(table.num_rows, remaining)
                frame = (
                    pa.concat_tables(tables).to_pandas()
                    if tables
                    else pd.DataFrame(columns=parquet.schema_arrow.names)
                )
            elif kind == "json":
                try:
                    frame = pd.read_json(path, orient="records")
                except ValueError:
                    frame = pd.read_json(path, lines=True)
                if len(frame) > self.max_rows:
                    raise SourceSafetyError(
                        f"JSON source exceeds the {self.max_rows:,} row safety limit"
                    )
                frame = frame.head(limit)
            else:
                raise ValueError(f"Unsupported local tabular source kind: {kind}")
        except (SourceSafetyError, SourceReadError):
            raise
        except (OSError, UnicodeError, ValueError, pa.ArrowException) as exc:
            raise SourceReadError(f"{kind.upper()} source is malformed or unreadable") from exc
        return self._validate_frame(frame)

    def _read_database(self, source: Mapping[str, Any], *, limit: int) -> pd.DataFrame:
        kind = str(source.get("kind", ""))
        table = source.get("table")
        if not isinstance(table, str) or not table:
            available = self.list_database_tables(source)
            suffix = f" Available tables: {', '.join(available[:20])}" if available else ""
            raise SourceReadError(f"A database table selection is required.{suffix}")
        quoted_table = _quote_identifier_path(table)
        try:
            available = self.list_database_tables(source)
            available_names = set(available)
            if kind == "postgresql":
                available_names |= {
                    value.removeprefix("public.")
                    for value in available
                    if value.startswith("public.")
                }
            if table not in available_names:
                raise SourceSafetyError("Selected database object is not an available base table")
            if kind == "sqlite":
                path = self._source_path(source)
                connection = sqlite3.connect(f"file:{path}?mode=ro", uri=True)
                try:
                    frame = pd.read_sql_query(
                        f"SELECT * FROM {quoted_table} LIMIT ?",  # noqa: S608 -- identifier validated
                        connection,
                        params=(int(limit),),
                    )
                finally:
                    connection.close()
            elif kind == "duckdb":
                path = self._source_path(source)
                connection = duckdb.connect(
                    str(path), read_only=True, config={"enable_external_access": "false"}
                )
                try:
                    frame = connection.execute(
                        f"SELECT * FROM {quoted_table} LIMIT ?",  # noqa: S608 -- identifier validated
                        [int(limit)],
                    ).fetch_df()
                finally:
                    connection.close()
            elif kind == "postgresql":
                engine = self._postgresql_engine(source)
                try:
                    frame = pd.read_sql_query(
                        text(f"SELECT * FROM {quoted_table} LIMIT :row_limit"),
                        engine,
                        params={"row_limit": int(limit)},
                    )
                finally:
                    engine.dispose()
            else:
                raise ValueError(f"Unsupported database source kind: {kind}")
        except (SourceSafetyError, SourceReadError):
            raise
        except (sqlite3.DatabaseError, duckdb.Error, OSError, ValueError) as exc:
            raise SourceReadError(f"{kind.upper()} source table could not be read") from exc
        except Exception as exc:
            if kind == "postgresql":
                raise SourceReadError(
                    "PostgreSQL source table could not be read using its environment reference"
                ) from exc
            raise
        return self._validate_frame(frame)

    def _read_source(
        self,
        source: Mapping[str, Any],
        *,
        limit: int | None = None,
    ) -> pd.DataFrame:
        bounded_limit = self.max_rows + 1 if limit is None else int(limit)
        if bounded_limit < 1 or bounded_limit > self.max_rows + 1:
            raise ValueError("Requested source row limit is outside the governed boundary")
        kind = str(source.get("kind", ""))
        if kind in {"csv", "xlsx", "parquet", "json"}:
            frame = self._read_local_tabular(source, limit=bounded_limit)
        elif kind in DATABASE_KINDS:
            frame = self._read_database(source, limit=bounded_limit)
        else:
            raise SourceSafetyError(f"Unsupported onboarding source kind: {kind or '(missing)'}")
        if len(frame) > self.max_rows:
            raise SourceSafetyError(f"Source exceeds the {self.max_rows:,} row safety limit")
        if frame.empty:
            raise SourceReadError("Source contains no data rows")
        return frame

    def preview_source(
        self,
        source: Mapping[str, Any],
        *,
        sample_rows: int | None = None,
    ) -> dict[str, Any]:
        """Preview bounded rows, inferred types and suggested exact-name mappings."""

        requested = self.max_preview_rows if sample_rows is None else int(sample_rows)
        sample_size = max(1, min(requested, self.max_preview_rows))
        frame = self._read_source(source, limit=sample_size)
        columns = []
        for name in frame.columns:
            inferred_type, confidence = _infer_series_type(frame[name])
            columns.append(
                {
                    "name": str(name),
                    "inferred_type": inferred_type,
                    "confidence": round(confidence, 6),
                    "null_count_in_sample": int(frame[name].isna().sum()),
                    "distinct_count_in_sample": int(frame[name].nunique(dropna=True)),
                }
            )
        normalised_source_names = {
            re.sub(r"[^a-z0-9]+", "_", str(name).casefold()).strip("_"): str(name)
            for name in frame.columns
        }
        suggested_mappings: dict[str, dict[str, str]] = {}
        for contract_id, contract in CONTRACTS.items():
            matches = {
                field.name: normalised_source_names[field.name]
                for field in contract.fields
                if field.name in normalised_source_names
            }
            suggested_mappings[contract_id] = matches
        return {
            "source": self._public_source(source),
            "sample_row_count": len(frame),
            "sample_limit": sample_size,
            "columns": columns,
            "rows": _portable_records(frame),
            "suggested_mappings": suggested_mappings,
        }

    @staticmethod
    def _public_source(source: Mapping[str, Any]) -> dict[str, Any]:
        allowed = {
            "source_id",
            "kind",
            "display_name",
            "relative_path",
            "size_bytes",
            "sha256",
            "table",
            "sheet",
            "url_env",
        }
        return {key: value for key, value in source.items() if key in allowed}

    @staticmethod
    def _validate_configuration(
        contract_id: str,
        mapping: Mapping[str, str],
        transformations: Mapping[str, str],
        source_columns: Sequence[str],
    ) -> tuple[DataContract, dict[str, SafeFormula]]:
        contract = CONTRACTS.get(contract_id)
        if contract is None:
            raise ValueError(f"Unknown canonical contract: {contract_id}")
        if not isinstance(mapping, Mapping) or not isinstance(transformations, Mapping):
            raise ValueError("mapping and transformations must be objects")
        fields = contract.field_lookup
        source_names = set(source_columns)
        unknown_targets = (set(mapping) | set(transformations)) - set(fields)
        if unknown_targets:
            raise ValueError(
                f"Unknown fields for {contract_id}: {', '.join(sorted(unknown_targets))}"
            )
        for target, source_name in mapping.items():
            if not isinstance(target, str) or not isinstance(source_name, str):
                raise ValueError("Mapping keys and values must be strings")
            if source_name not in source_names:
                raise ValueError(f"Mapped source field does not exist: {source_name}")
        provided = set(mapping) | set(transformations)
        missing = [
            field.name for field in contract.fields if field.required and field.name not in provided
        ]
        if missing:
            raise ValueError(f"Required contract mappings are missing: {', '.join(missing)}")
        transformation_targets = set(transformations)
        dependencies: dict[str, set[str]] = {}
        for target, expression in transformations.items():
            if not isinstance(target, str) or not isinstance(expression, str):
                raise ValueError("Transformation keys and expressions must be strings")
            try:
                parsed = ast.parse(expression, mode="eval")
            except SyntaxError as exc:
                raise FormulaSafetyError("Formula syntax is invalid") from exc
            names = {
                node.id
                for node in ast.walk(parsed)
                if isinstance(node, ast.Name) and node.id not in SafeFormula._allowed_functions
            }
            if target in names and target not in source_names and target not in mapping:
                raise FormulaSafetyError(f"Transformation {target} depends on itself")
            dependencies[target] = (names & transformation_targets) - {target}
        ordered_targets: list[str] = []
        remaining = list(transformations)
        while remaining:
            ready = [target for target in remaining if dependencies[target] <= set(ordered_targets)]
            if not ready:
                raise FormulaSafetyError("Derived-field transformations contain a dependency cycle")
            for target in ready:
                ordered_targets.append(target)
                remaining.remove(target)
        allowed_formula_fields = set(source_names) | set(mapping)
        compiled: dict[str, SafeFormula] = {}
        for target in ordered_targets:
            expression = transformations[target]
            compiled[target] = SafeFormula(
                expression,
                allowed_fields=sorted(allowed_formula_fields),
            )
            allowed_formula_fields.add(target)
        return contract, compiled

    def _canonicalise(
        self,
        source_frame: pd.DataFrame,
        *,
        contract: DataContract,
        mapping: Mapping[str, str],
        transformations: Mapping[str, SafeFormula],
    ) -> tuple[pd.DataFrame, dict[int, list[str]], tuple[dict[str, Any], ...], int]:
        frame = source_frame.reset_index(drop=True)
        canonical = pd.DataFrame(index=frame.index)
        for field in contract.fields:
            canonical[field.name] = pd.Series([None] * len(frame), dtype="object")
        for target, source_name in mapping.items():
            canonical[target] = frame[source_name].copy()

        row_errors: dict[int, list[str]] = {}
        error_preview: list[dict[str, Any]] = []
        error_count = 0

        def record_error(
            row_index: int,
            field: str,
            code: str,
            message: str,
            value: Any,
        ) -> None:
            nonlocal error_count
            error_count += 1
            row_errors.setdefault(row_index, []).append(f"{field}:{code}")
            if len(error_preview) < self.max_error_preview:
                error_preview.append(
                    {
                        "row_number": row_index + 1,
                        "field": field,
                        "code": code,
                        "message": message,
                        "value": _portable_value(value),
                    }
                )

        if transformations:
            records = frame.to_dict(orient="records")
            for row_index, raw_row in enumerate(records):
                context = dict(raw_row)
                context.update({target: canonical.at[row_index, target] for target in mapping})
                for target, formula in transformations.items():
                    try:
                        value = formula.evaluate(context)
                    except FormulaSafetyError as exc:
                        value = None
                        record_error(
                            row_index,
                            target,
                            "TRANSFORMATION_ERROR",
                            str(exc),
                            None,
                        )
                    canonical.at[row_index, target] = value
                    context[target] = value

        for field in contract.fields:
            raw = canonical[field.name].copy()
            invalid = pd.Series(False, index=canonical.index)
            if field.data_type == "string":
                converted = raw.map(lambda value: None if _is_null(value) else str(value).strip())
                converted = converted.map(lambda value: None if value == "" else value)
            elif field.data_type in {"integer", "number"}:
                numeric = pd.to_numeric(raw, errors="coerce")
                invalid = raw.notna() & numeric.isna()
                non_finite = numeric.notna() & ~numeric.map(
                    lambda value: math.isfinite(float(value))
                )
                invalid |= non_finite
                numeric = numeric.mask(non_finite)
                if field.data_type == "integer":
                    fractional = numeric.notna() & ((numeric % 1).abs() > 1e-12)
                    invalid |= fractional
                    numeric = numeric.mask(fractional)
                    converted = numeric.astype("Int64")
                else:
                    converted = numeric.astype("Float64")
            elif field.data_type in {"date", "datetime"}:
                parsed = pd.to_datetime(raw, errors="coerce", utc=True)
                invalid = raw.notna() & parsed.isna()
                converted = parsed.dt.date if field.data_type == "date" else parsed
            elif field.data_type == "boolean":
                converted, invalid = _boolean_series(raw)
            else:  # pragma: no cover - contracts are defined in this module
                raise AssertionError(f"Unsupported contract type: {field.data_type}")
            canonical[field.name] = converted
            for row_index in canonical.index[invalid]:
                record_error(
                    int(row_index),
                    field.name,
                    "INVALID_TYPE",
                    f"Expected {field.data_type}",
                    raw.at[row_index],
                )
            if field.required:
                required_invalid = canonical[field.name].isna() & ~invalid
                for row_index in canonical.index[required_invalid]:
                    record_error(
                        int(row_index),
                        field.name,
                        "REQUIRED",
                        "Required value is missing",
                        raw.at[row_index],
                    )
            if field.non_negative:
                negative = canonical[field.name].notna() & (canonical[field.name] < 0)
                for row_index in canonical.index[negative]:
                    record_error(
                        int(row_index),
                        field.name,
                        "NEGATIVE_VALUE",
                        "Value must be non-negative",
                        canonical.at[row_index, field.name],
                    )
            if field.allowed_values:
                disallowed = canonical[field.name].notna() & ~canonical[field.name].isin(
                    field.allowed_values
                )
                for row_index in canonical.index[disallowed]:
                    record_error(
                        int(row_index),
                        field.name,
                        "DISALLOWED_VALUE",
                        f"Value must be one of: {', '.join(field.allowed_values)}",
                        canonical.at[row_index, field.name],
                    )

        if all(key in canonical for key in contract.unique_key):
            complete_key = canonical[list(contract.unique_key)].notna().all(axis=1)
            duplicates = complete_key & canonical.duplicated(list(contract.unique_key), keep=False)
            for row_index in canonical.index[duplicates]:
                record_error(
                    int(row_index),
                    ",".join(contract.unique_key),
                    "DUPLICATE_KEY",
                    "Canonical unique key is duplicated",
                    "|".join(
                        str(_portable_value(canonical.at[row_index, key]))
                        for key in contract.unique_key
                    ),
                )

        date_pairs = {
            "account_master": ("opened_date", "closed_date"),
            "membership_history": ("start_date", "end_date"),
        }
        if contract.contract_id in date_pairs:
            start_name, end_name = date_pairs[contract.contract_id]
            reversed_dates = (
                canonical[start_name].notna()
                & canonical[end_name].notna()
                & (canonical[end_name] < canonical[start_name])
            )
            for row_index in canonical.index[reversed_dates]:
                record_error(
                    int(row_index),
                    end_name,
                    "INVALID_DATE_RANGE",
                    f"{end_name} cannot precede {start_name}",
                    canonical.at[row_index, end_name],
                )
        return canonical, row_errors, tuple(error_preview), error_count

    def _validation_frames(
        self,
        frame: pd.DataFrame,
        *,
        contract_id: str,
        mapping: Mapping[str, str],
        transformations: Mapping[str, str],
        max_error_rate: float,
    ) -> _ValidationFrames:
        threshold = float(max_error_rate)
        if not 0 <= threshold <= 1:
            raise ValueError("max_error_rate must be between 0 and 1")
        contract, compiled = self._validate_configuration(
            contract_id,
            mapping,
            transformations,
            [str(column) for column in frame.columns],
        )
        canonical, row_errors, error_preview, error_count = self._canonicalise(
            frame,
            contract=contract,
            mapping=mapping,
            transformations=compiled,
        )
        invalid_indexes = sorted(row_errors)
        valid_mask = ~canonical.index.isin(invalid_indexes)
        valid = canonical.loc[valid_mask].reset_index(drop=True)
        invalid = canonical.loc[~valid_mask].copy()
        if not invalid.empty:
            invalid.insert(0, "_source_row_number", [int(index) + 1 for index in invalid.index])
            invalid["_error_codes"] = [";".join(row_errors[int(index)]) for index in invalid.index]
        else:
            invalid.insert(0, "_source_row_number", pd.Series(dtype="int64"))
            invalid["_error_codes"] = pd.Series(dtype="string")
        invalid = invalid.reset_index(drop=True)
        source_rows = len(canonical)
        invalid_rows = len(invalid_indexes)
        error_rate = invalid_rows / source_rows if source_rows else 0.0
        summary = {
            "contract_id": contract.contract_id,
            "contract_version": contract.version,
            "source_rows": source_rows,
            "valid_rows": len(valid),
            "invalid_rows": invalid_rows,
            "validation_error_count": error_count,
            "error_preview_count": len(error_preview),
            "error_preview_truncated": error_count > len(error_preview),
            "error_rate": round(error_rate, 10),
            "max_error_rate": threshold,
            "passed": error_rate <= threshold,
        }
        return _ValidationFrames(
            canonical=canonical,
            valid=valid,
            invalid=invalid,
            source_row_numbers=pd.Series(range(1, source_rows + 1)),
            errors=error_preview,
            error_rows=frozenset(invalid_indexes),
            summary=summary,
        )

    def validate_source(
        self,
        source: Mapping[str, Any],
        *,
        contract_id: str,
        mapping: Mapping[str, str],
        transformations: Mapping[str, str] | None = None,
        max_error_rate: float = 0.0,
    ) -> dict[str, Any]:
        """Validate without writing rows or changing profile state."""

        frame = self._read_source(source)
        validated = self._validation_frames(
            frame,
            contract_id=contract_id,
            mapping=mapping,
            transformations=transformations or {},
            max_error_rate=max_error_rate,
        )
        return {
            "source": self._public_source(source),
            "validation": validated.summary,
            "error_preview": list(validated.errors),
            "valid_row_preview": _portable_records(validated.valid, self.max_preview_rows),
            "invalid_row_preview": _portable_records(validated.invalid, self.max_preview_rows),
        }

    def validate_mapping(
        self,
        source: Mapping[str, Any],
        *,
        contract_id: str,
        mapping: Mapping[str, str],
        transformations: Mapping[str, str] | None = None,
    ) -> dict[str, Any]:
        """Validate mapping and formula syntax against a bounded source schema."""

        frame = self._read_source(source, limit=1)
        contract, compiled = self._validate_configuration(
            contract_id,
            mapping,
            transformations or {},
            [str(column) for column in frame.columns],
        )
        return {
            "valid": True,
            "contract_id": contract.contract_id,
            "contract_version": contract.version,
            "mapped_fields": sorted(mapping),
            "derived_fields": sorted(compiled),
            "source_fields_used": sorted(set(mapping.values())),
        }

    def _profile_path(self, profile_id: str) -> Path:
        return self.profiles_root / f"{_profile_identifier(profile_id)}.json"

    @staticmethod
    def _external_id(profile_id: str) -> str:
        return f"{PROFILE_EXTERNAL_PREFIX}{_profile_identifier(profile_id)}"

    @staticmethod
    def _profile_response(record: Mapping[str, Any]) -> dict[str, Any]:
        state = dict(record["state"])
        state.update(
            {
                "version": int(record["version"]),
                "approval_state": str(record["approval_state"]),
                "created_at": record.get("created_at"),
                "modified_at": record.get("modified_at"),
                "created_by": record.get("created_by"),
                "modified_by": record.get("modified_by"),
            }
        )
        return state

    def _write_profile_record(self, profile_id: str, record: Mapping[str, Any]) -> None:
        portable_record = {
            "schema_version": "1.0.0",
            "object_type": PROFILE_OBJECT_TYPE,
            "external_id": self._external_id(profile_id),
            "version": int(record["version"]),
            "approval_state": str(record["approval_state"]),
            "created_at": record.get("created_at"),
            "modified_at": record.get("modified_at"),
            "created_by": record.get("created_by"),
            "modified_by": record.get("modified_by"),
            "state": dict(record["state"]),
        }
        _atomic_json(self._profile_path(profile_id), portable_record)

    def _get_profile_record(self, profile_id: str) -> dict[str, Any]:
        profile_id = _profile_identifier(profile_id)
        if self.workflow_store is not None:
            try:
                return self.workflow_store.get(
                    PROFILE_OBJECT_TYPE,
                    self._external_id(profile_id),
                )
            except ObjectNotFound as exc:
                raise OnboardingError(f"Onboarding profile not found: {profile_id}") from exc
        path = self._profile_path(profile_id)
        if not path.is_file():
            raise OnboardingError(f"Onboarding profile not found: {profile_id}")
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError) as exc:
            raise OnboardingError(f"Onboarding profile is unreadable: {profile_id}") from exc
        if payload.get("external_id") != self._external_id(profile_id):
            raise SourceSafetyError("Onboarding profile identifier does not match its filename")
        return payload

    def get_profile(self, profile_id: str) -> dict[str, Any]:
        """Return one reusable profile without exposing source credentials."""

        return self._profile_response(self._get_profile_record(profile_id))

    def list_profiles(self) -> list[dict[str, Any]]:
        """List persisted profiles and their current approval state."""

        if self.workflow_store is not None:
            records = [
                record
                for record in self.workflow_store.list(PROFILE_OBJECT_TYPE)
                if str(record["external_id"]).startswith(PROFILE_EXTERNAL_PREFIX)
            ]
        else:
            records = []
            for path in sorted(self.profiles_root.glob("*.json")):
                try:
                    records.append(json.loads(path.read_text(encoding="utf-8")))
                except (OSError, json.JSONDecodeError):
                    continue
        return [self._profile_response(record) for record in records]

    def save_import_profile(
        self,
        profile_id: str,
        source: Mapping[str, Any],
        *,
        contract_id: str,
        mapping: Mapping[str, str],
        transformations: Mapping[str, str] | None = None,
        max_error_rate: float = 0.0,
        actor: str,
    ) -> dict[str, Any]:
        """Validate and persist a draft import profile for future compatible sources."""

        profile_id = _profile_identifier(profile_id)
        if not actor.strip() or len(actor) > 128:
            raise ValueError("actor is required and may contain at most 128 characters")
        transformation_payload = dict(transformations or {})
        frame = self._read_source(source)
        validated = self._validation_frames(
            frame,
            contract_id=contract_id,
            mapping=mapping,
            transformations=transformation_payload,
            max_error_rate=max_error_rate,
        )
        contract = CONTRACTS[contract_id]
        source_schema = []
        for column in frame.columns:
            inferred_type, confidence = _infer_series_type(frame[column])
            source_schema.append(
                {
                    "name": str(column),
                    "inferred_type": inferred_type,
                    "confidence": round(confidence, 6),
                }
            )
        now = datetime.now(UTC).isoformat()
        _, ordered_formulas = self._validate_configuration(
            contract_id,
            mapping,
            transformation_payload,
            [str(column) for column in frame.columns],
        )
        required_source_columns = set(mapping.values())
        overridden_context = set(mapping)
        for target, formula in ordered_formulas.items():
            expression = formula.expression
            referenced_names = {
                node.id
                for node in ast.walk(ast.parse(expression, mode="eval"))
                if isinstance(node, ast.Name) and node.id not in SafeFormula._allowed_functions
            }
            required_source_columns |= {
                name
                for name in referenced_names
                if name in frame.columns and name not in overridden_context
            }
            overridden_context.add(target)
        state = {
            "schema_version": "1.0.0",
            "profile_id": profile_id,
            "contract_id": contract.contract_id,
            "contract_version": contract.version,
            "mapping": dict(mapping),
            "transformations": transformation_payload,
            "max_error_rate": float(max_error_rate),
            "required_source_columns": sorted(required_source_columns),
            "source_schema": source_schema,
            "original_source": self._public_source(source),
            "draft_validation": validated.summary,
            "active": False,
            "last_run": None,
            "created_at_utc": now,
            "updated_at_utc": now,
        }
        if self.workflow_store is not None:
            try:
                record = self.workflow_store.create(
                    PROFILE_OBJECT_TYPE,
                    self._external_id(profile_id),
                    state,
                    actor=actor,
                    approval_state="DRAFT",
                )
            except DuplicateObject as exc:
                raise OnboardingError(f"Onboarding profile already exists: {profile_id}") from exc
        else:
            if self._profile_path(profile_id).exists():
                raise OnboardingError(f"Onboarding profile already exists: {profile_id}")
            record = {
                "version": 1,
                "approval_state": "DRAFT",
                "created_at": now,
                "modified_at": now,
                "created_by": actor,
                "modified_by": actor,
                "state": state,
            }
        self._write_profile_record(profile_id, record)
        response = self._profile_response(record)
        response["error_preview"] = list(validated.errors)
        return response

    @staticmethod
    def _numeric_total(series: pd.Series) -> float | None:
        numeric = pd.to_numeric(series, errors="coerce")
        finite = numeric[np.isfinite(numeric)]
        return float(finite.sum()) if not finite.empty else None

    @staticmethod
    def _relative_output(root: Path, path: Path) -> str:
        return path.resolve().relative_to(root.resolve()).as_posix()

    @staticmethod
    def _write_parquet(frame: pd.DataFrame, path: Path) -> None:
        path.parent.mkdir(parents=True, exist_ok=False)
        table = pa.Table.from_pandas(frame, preserve_index=False)
        pq.write_table(table, path, compression="zstd")
        path.chmod(0o640)

    def _reconcile(
        self,
        validated: _ValidationFrames,
        contract: DataContract,
    ) -> dict[str, Any]:
        numeric_fields = [
            field.name for field in contract.fields if field.data_type in {"integer", "number"}
        ]
        totals: list[dict[str, Any]] = []
        balanced = True
        for field in numeric_fields:
            source_total = self._numeric_total(validated.canonical[field])
            loaded_total = self._numeric_total(validated.valid[field])
            quarantined_total = self._numeric_total(validated.invalid[field])
            if source_total is None:
                delta = None
                field_balanced = True
            else:
                loaded_value = loaded_total or 0.0
                quarantine_value = quarantined_total or 0.0
                delta = source_total - loaded_value - quarantine_value
                tolerance = max(1e-8, abs(source_total) * 1e-10)
                field_balanced = abs(delta) <= tolerance
            balanced = balanced and field_balanced
            totals.append(
                {
                    "field": field,
                    "source_total": source_total,
                    "loaded_total": loaded_total,
                    "quarantined_total": quarantined_total,
                    "balance_delta": delta,
                    "balanced": field_balanced,
                }
            )
        source_rows = int(validated.summary["source_rows"])
        loaded_rows = int(validated.summary["valid_rows"])
        quarantined_rows = int(validated.summary["invalid_rows"])
        row_delta = source_rows - loaded_rows - quarantined_rows
        return {
            "source_rows": source_rows,
            "loaded_rows": loaded_rows,
            "quarantined_rows": quarantined_rows,
            "row_balance_delta": row_delta,
            "numeric_totals": totals,
            "balanced": balanced and row_delta == 0,
        }

    def _preview_mart_frame(
        self,
        validated: _ValidationFrames,
        reconciliation: Mapping[str, Any],
    ) -> pd.DataFrame:
        metrics: list[dict[str, Any]] = [
            {"metric": "source_rows", "value": reconciliation["source_rows"]},
            {"metric": "loaded_rows", "value": reconciliation["loaded_rows"]},
            {"metric": "quarantined_rows", "value": reconciliation["quarantined_rows"]},
            {"metric": "error_rate", "value": validated.summary["error_rate"]},
        ]
        for item in reconciliation["numeric_totals"]:
            metrics.extend(
                [
                    {"metric": f"{item['field']}.source_total", "value": item["source_total"]},
                    {"metric": f"{item['field']}.loaded_total", "value": item["loaded_total"]},
                    {
                        "metric": f"{item['field']}.quarantined_total",
                        "value": item["quarantined_total"],
                    },
                ]
            )
        return pd.DataFrame(metrics)

    def run_import_profile(
        self,
        profile_id: str,
        source: Mapping[str, Any],
        *,
        actor: str,
        expected_version: int | None = None,
    ) -> dict[str, Any]:
        """Load a compatible source into quarantine-isolated onboarding outputs."""

        profile_id = _profile_identifier(profile_id)
        if not actor.strip() or len(actor) > 128:
            raise ValueError("actor is required and may contain at most 128 characters")
        current = self._get_profile_record(profile_id)
        current_version = int(current["version"])
        if expected_version is not None and int(expected_version) != current_version:
            raise ConcurrencyConflict(
                f"Expected onboarding profile {profile_id} version {expected_version}"
            )
        state = dict(current["state"])
        contract = CONTRACTS.get(str(state.get("contract_id")))
        if contract is None or state.get("contract_version") != contract.version:
            raise OnboardingError("Import profile references an unavailable contract version")
        frame = self._read_source(source)
        missing_columns = set(state["required_source_columns"]) - set(frame.columns)
        if missing_columns:
            raise OnboardingError(
                f"Source is incompatible with the profile; missing: {', '.join(sorted(missing_columns))}"
            )
        validated = self._validation_frames(
            frame,
            contract_id=contract.contract_id,
            mapping=state["mapping"],
            transformations=state["transformations"],
            max_error_rate=float(state["max_error_rate"]),
        )
        reconciliation = self._reconcile(validated, contract)
        run_id = f"{datetime.now(UTC).strftime('%Y%m%dT%H%M%SZ')}-{uuid.uuid4().hex[:12]}"
        namespace_path = self.namespace_root / profile_id / run_id / "data.parquet"
        quarantine_path = self.quarantine_root / profile_id / run_id / "invalid_rows.parquet"
        preview_path = self.preview_marts_root / profile_id / run_id / "summary.parquet"
        preview_manifest_path = self.preview_marts_root / profile_id / run_id / "run.json"
        created_directories = [namespace_path.parent, quarantine_path.parent, preview_path.parent]
        try:
            self._write_parquet(validated.valid, namespace_path)
            self._write_parquet(validated.invalid, quarantine_path)
            self._write_parquet(
                self._preview_mart_frame(validated, reconciliation),
                preview_path,
            )
            run_result = {
                "schema_version": "1.0.0",
                "run_id": run_id,
                "profile_id": profile_id,
                "profile_version": current_version,
                "contract_id": contract.contract_id,
                "contract_version": contract.version,
                "source": self._public_source(source),
                "validation": validated.summary,
                "error_preview": list(validated.errors),
                "reconciliation": reconciliation,
                "outputs": {
                    "onboarding_namespace": self._relative_output(self.root, namespace_path),
                    "quarantine": self._relative_output(self.root, quarantine_path),
                    "preview_mart": self._relative_output(self.root, preview_path),
                },
                "output_hashes": {
                    "onboarding_namespace": _sha256(namespace_path),
                    "quarantine": _sha256(quarantine_path),
                    "preview_mart": _sha256(preview_path),
                },
                "loaded_to_active_analytics": False,
                "ran_at_utc": datetime.now(UTC).isoformat(),
                "actor": actor,
            }
            _atomic_json(preview_manifest_path, run_result)
        except Exception:
            for directory in created_directories:
                shutil.rmtree(directory, ignore_errors=True)
            raise
        state["last_run"] = run_result
        state["updated_at_utc"] = datetime.now(UTC).isoformat()
        if self.workflow_store is not None:
            updated = self.workflow_store.update(
                PROFILE_OBJECT_TYPE,
                self._external_id(profile_id),
                state,
                expected_version=current_version,
                actor=actor,
                replace=True,
            )
        else:
            now = datetime.now(UTC).isoformat()
            updated = {
                **current,
                "version": current_version + 1,
                "modified_at": now,
                "modified_by": actor,
                "state": state,
            }
        self._write_profile_record(profile_id, updated)
        return {
            **run_result,
            "profile_version": int(updated["version"]),
            "profile_approval_state": str(updated["approval_state"]),
            "profile_active": bool(updated["state"].get("active", False)),
        }

    def load_into_onboarding_namespace(
        self,
        profile_id: str,
        source: Mapping[str, Any],
        *,
        actor: str,
        expected_version: int | None = None,
    ) -> dict[str, Any]:
        """API-friendly alias for :meth:`run_import_profile`."""

        return self.run_import_profile(
            profile_id,
            source,
            actor=actor,
            expected_version=expected_version,
        )

    def approve_profile(
        self,
        profile_id: str,
        *,
        expected_version: int,
        actor: str,
        rationale: str,
    ) -> dict[str, Any]:
        """Activate a profile only after a passing, fully reconciled onboarding run."""

        profile_id = _profile_identifier(profile_id)
        if not actor.strip() or len(actor) > 128 or not rationale.strip() or len(rationale) > 2_000:
            raise ValueError(
                "actor (up to 128 characters) and approval rationale "
                "(up to 2,000 characters) are required"
            )
        current = self._get_profile_record(profile_id)
        current_version = int(current["version"])
        if current_version != int(expected_version):
            raise ConcurrencyConflict(
                f"Expected onboarding profile {profile_id} version {expected_version}"
            )
        state = dict(current["state"])
        last_run = state.get("last_run")
        if not isinstance(last_run, Mapping):
            raise ProfileApprovalError("Profile must complete an onboarding run before approval")
        if not bool(last_run.get("validation", {}).get("passed")):
            raise ProfileApprovalError("Profile's latest onboarding run did not pass validation")
        if not bool(last_run.get("reconciliation", {}).get("balanced")):
            raise ProfileApprovalError("Profile's latest onboarding run did not reconcile")
        state.update(
            {
                "active": True,
                "approved_at_utc": datetime.now(UTC).isoformat(),
                "approved_by": actor,
                "approval_rationale": rationale,
                "updated_at_utc": datetime.now(UTC).isoformat(),
            }
        )
        if self.workflow_store is not None:
            updated = self.workflow_store.update(
                PROFILE_OBJECT_TYPE,
                self._external_id(profile_id),
                state,
                expected_version=current_version,
                actor=actor,
                approval_state="APPROVED",
                replace=True,
            )
        else:
            now = datetime.now(UTC).isoformat()
            updated = {
                **current,
                "version": current_version + 1,
                "approval_state": "APPROVED",
                "modified_at": now,
                "modified_by": actor,
                "state": state,
            }
        self._write_profile_record(profile_id, updated)
        return self._profile_response(updated)


__all__ = [
    "CONTRACTS",
    "DataContract",
    "FormulaSafetyError",
    "OnboardingError",
    "OnboardingStudio",
    "ProfileApprovalError",
    "SafeFormula",
    "SourceReadError",
    "SourceSafetyError",
    "list_contracts",
]
