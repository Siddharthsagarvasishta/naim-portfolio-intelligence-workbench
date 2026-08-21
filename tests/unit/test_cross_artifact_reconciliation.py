from __future__ import annotations

import json
import subprocess
from pathlib import Path

from openpyxl import Workbook

from scripts.reconcile_release_artifacts import (
    FAIL,
    INCOMPLETE,
    MISSING,
    PASS,
    _linkedin_channel,
    _parse_hyper_shell_rows,
    _pdf_text_with_pdfkit,
    _release_manifest_checks,
    _snapshot_story,
    _workbook_channel,
    channel,
    check,
    compare_story,
    extract_canonical_story,
    sha256_file,
    verify_file_ledger,
)


def _story() -> dict[str, object]:
    return {
        "metric_id": "ANNUALISED_NET_LOSS_RATE",
        "metric_registry_version": "1.0.0",
        "metric_version": "1.0.0",
        "reporting_period": "2025-08-01",
        "comparison_period": "2025-07-01",
        "scope": "All portfolio",
        "current_annualised_net_loss_rate": 0.06685632988073756,
        "prior_annualised_net_loss_rate": 0.035714829388391336,
        "observed_change_bps": 311.4150049234624,
        "mix_contribution_bps": 4.433506460154617,
        "within_segment_contribution_bps": 306.9814984633076,
        "reconciliation_residual_bps": 5.204170427930421e-14,
        "primary_dimension": "acquisition_channel",
        "primary_driver": "Affiliate",
        "causal_status": "ASSOCIATIONAL",
        "data_quality_status": "PASS",
        "publication_allowed": True,
        "synthetic_data": True,
        "run_id": "default-73421-6006e471387a",
        "configuration_hash": "c" * 64,
        "dataset_hash": "d" * 64,
        "evidence_id": "NAIM-default-73421-6006e471387a",
        "evidence_payload_sha256": "e" * 64,
        "row_count": 513923,
        "canonical_file_sha256": "f" * 64,
    }


def _canonical_snapshot() -> dict[str, object]:
    return {
        "schema_version": "1.0.0",
        "evidence_id": "NAIM-default-73421-6006e471387a",
        "evidence_sha256": "placeholder",
        "synthetic_data_flag": True,
        "metadata": {
            "run_id": "default-73421-6006e471387a",
            "configuration_hash": "c" * 64,
            "metric_registry_version": "1.0.0",
            "row_counts": {"monthly_account_performance": 513923},
        },
        "kpis": [
            {
                "metric_id": "ANNUALISED_NET_LOSS_RATE",
                "metric_version": "1.0.0",
                "reporting_period": "2025-08-01",
                "comparison_period": "2025-07-01",
                "value": 0.06685632988073756,
                "prior_value": 0.035714829388391336,
            }
        ],
        "root_cause": {
            "finding": {
                "observed_change_bps": 311.4150049234624,
                "mix_contribution_bps": 4.433506460154617,
                "within_segment_contribution_bps": 306.9814984633076,
                "reconciliation_residual_bps": 5.204170427930421e-14,
                "primary_dimension": "acquisition_channel",
                "primary_driver": "Affiliate",
                "causal_status": "ASSOCIATIONAL",
            }
        },
        "data_quality": {"status": "PASS", "publication_allowed": True},
    }


def _write_release_manifest(
    repository_root: Path,
    artifact: Path,
    expected: dict[str, object],
    *,
    filter_scope: object = "All portfolio",
) -> None:
    target = repository_root / "outputs" / "manifests" / "artifact.manifest.json"
    target.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "artifact_id": "ART-001",
        "built_at_utc": "2026-08-01T00:00:00+00:00",
        "artifact": {
            "path": artifact.relative_to(repository_root).as_posix(),
            "sha256": sha256_file(artifact),
        },
        "versions": {
            "dataset_hash": {"value": expected["dataset_hash"]},
            "configuration_hash": {"value": expected["configuration_hash"]},
            "script_version": {"value": "1.0.0"},
        },
        "metric_registry_version": expected["metric_registry_version"],
        "reporting_period": expected["reporting_period"],
        "filter_scope": filter_scope,
        "evidence_ids": [expected["evidence_id"]],
        "data_quality_result": expected["data_quality_status"],
        "synthetic_data": expected["synthetic_data"],
    }
    target.write_text(json.dumps(payload), encoding="utf-8")


def test_extract_canonical_story_preserves_full_precision() -> None:
    story = extract_canonical_story(_canonical_snapshot())

    assert story["observed_change_bps"] == 311.4150049234624
    assert story["mix_contribution_bps"] == 4.433506460154617
    assert story["within_segment_contribution_bps"] == 306.9814984633076
    assert story["primary_driver"] == "Affiliate"
    assert story["causal_status"] == "ASSOCIATIONAL"


def test_numeric_comparison_is_tolerant_only_within_declared_bound() -> None:
    accepted = check("metric", expected=100.0, actual=100.0 + 9e-10, tolerance=1e-9)
    rejected = check("metric", expected=100.0, actual=100.0 + 2e-9, tolerance=1e-9)

    assert accepted["outcome"] == PASS
    assert rejected["outcome"] == FAIL


def test_compare_story_marks_absent_required_claim_as_missing() -> None:
    actual = dict(_story())
    actual.pop("primary_driver")

    checks = compare_story(_story(), actual, prefix="sample", fields=["primary_driver"])

    assert checks == [
        {
            "check_id": "sample.story.primary_driver",
            "required": True,
            "outcome": MISSING,
            "expected": "Affiliate",
            "actual": None,
        }
    ]


def test_channel_is_incomplete_for_missing_evidence_and_fails_for_mismatch() -> None:
    incomplete = channel("a", "A", checks=[check("missing")])
    failed = channel("b", "B", checks=[check("mismatch", expected=1, actual=2)])

    assert incomplete["status"] == INCOMPLETE
    assert failed["status"] == FAIL


def test_file_ledger_rehashes_files_and_detects_tampering(tmp_path: Path) -> None:
    root = tmp_path / "package"
    root.mkdir()
    artifact = root / "evidence.txt"
    artifact.write_text("governed", encoding="utf-8")
    ledger = [
        {
            "path": "evidence.txt",
            "bytes": artifact.stat().st_size,
            "sha256": sha256_file(artifact),
        }
    ]

    passing = verify_file_ledger(root, ledger, prefix="package", repository_root=tmp_path)
    artifact.write_text("tampered", encoding="utf-8")
    failing = verify_file_ledger(root, ledger, prefix="package", repository_root=tmp_path)

    assert all(item["outcome"] == PASS for item in passing)
    assert any(item["check_id"].endswith("sha256") and item["outcome"] == FAIL for item in failing)


def test_pdfkit_fallback_passes_pdf_path_and_returns_native_text(
    tmp_path: Path, monkeypatch
) -> None:
    pdf = tmp_path / "carousel.pdf"
    pdf.write_bytes(b"%PDF-1.7\n")
    captured: list[list[str]] = []

    monkeypatch.setattr("scripts.reconcile_release_artifacts.sys.platform", "darwin")
    monkeypatch.setattr(
        "scripts.reconcile_release_artifacts.shutil.which",
        lambda name: "/usr/bin/swift" if name == "swift" else None,
    )

    def fake_run(command, **kwargs):
        captured.append(command)
        assert kwargs == {"check": True, "capture_output": True, "text": True}
        return subprocess.CompletedProcess(command, 0, stdout="Name the movement.\n")

    monkeypatch.setattr("scripts.reconcile_release_artifacts.subprocess.run", fake_run)

    assert _pdf_text_with_pdfkit(pdf) == "Name the movement.\n"
    assert captured[0][-1] == str(pdf)
    assert captured[0][0] == "/usr/bin/swift"


def test_hyper_shell_parser_requires_all_native_row_classes() -> None:
    output = """
NAIMPORTFOLIO|2025-07-01|10|1|100
NAIMPORTFOLIO|2025-08-01|20|2|100
NAIMSEGMENT|2025-07-01|Affiliate|10|1|100
NAIMSEGMENT|2025-08-01|Affiliate|20|2|100
NAIMMETA|product|nAIM Portfolio Intelligence Workbench
"""

    portfolio, segments, metadata = _parse_hyper_shell_rows(output)

    assert portfolio == [["2025-07-01", "10", "1", "100"], ["2025-08-01", "20", "2", "100"]]
    assert segments[-1][1] == "Affiliate"
    assert metadata == [["product", "nAIM Portfolio Intelligence Workbench"]]


def test_release_manifest_requires_and_validates_full_provenance(tmp_path: Path) -> None:
    artifact = tmp_path / "outputs" / "sample.html"
    artifact.parent.mkdir(parents=True)
    artifact.write_text("<html></html>", encoding="utf-8")
    expected = _story()
    _write_release_manifest(tmp_path, artifact, expected)

    checks, paths = _release_manifest_checks(tmp_path, artifact, expected, prefix="sample.manifest")

    assert paths == ["outputs/manifests/artifact.manifest.json"]
    assert all(item["outcome"] == PASS for item in checks)


def test_release_manifest_normalises_machine_readable_all_portfolio_scope(
    tmp_path: Path,
) -> None:
    artifact = tmp_path / "outputs" / "sample.html"
    artifact.parent.mkdir(parents=True)
    artifact.write_text("<html></html>", encoding="utf-8")
    expected = _story()
    _write_release_manifest(
        tmp_path,
        artifact,
        expected,
        filter_scope={"headline_scope": "all_portfolio"},
    )

    checks, _ = _release_manifest_checks(tmp_path, artifact, expected, prefix="sample.manifest")

    scope_check = next(item for item in checks if item["check_id"].endswith("filter_scope"))
    assert scope_check["outcome"] == PASS


def test_release_manifest_missing_is_never_a_pass(tmp_path: Path) -> None:
    artifact = tmp_path / "outputs" / "sample.html"
    artifact.parent.mkdir(parents=True)
    artifact.write_text("<html></html>", encoding="utf-8")

    checks, paths = _release_manifest_checks(tmp_path, artifact, _story(), prefix="sample.manifest")

    assert paths == []
    assert checks[0]["outcome"] == MISSING


def test_public_snapshot_adapter_maps_the_governed_story() -> None:
    payload = {
        "reporting_period": "2025-08-01",
        "evidence_id": "NAIM-default-73421-6006e471387a-2025-08",
        "source_context": {
            "run_id": "default-73421-6006e471387a",
            "metric_registry_version": "1.0.0",
        },
        "portfolio_story": {
            "metric_id": "ANNUALISED_NET_LOSS_RATE",
            "metric_version": "1.0.0",
            "current_annualised_net_loss_rate": 0.06685632988073756,
            "prior_annualised_net_loss_rate": 0.035714829388391336,
            "observed_change_bps": 311.4150049234624,
            "data_quality_status": "PASS",
        },
        "decomposition": {
            "dimension": "acquisition_channel",
            "primary_driver": "Affiliate",
            "causal_status": "ASSOCIATIONAL",
            "mix_bps": 4.433506460154617,
            "within_segment_bps": 306.9814984633076,
            "residual_bps": 5.204170427930421e-14,
        },
    }

    actual = _snapshot_story(payload)

    assert actual["observed_change_bps"] == _story()["observed_change_bps"]
    assert actual["primary_driver"] == "Affiliate"


def test_workbook_adapter_reads_native_cells_and_release_manifest(tmp_path: Path) -> None:
    expected = _story()
    path = tmp_path / "outputs" / "nAIM_Portfolio_Intelligence_Workbench.xlsx"
    path.parent.mkdir(parents=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    root = workbook.create_sheet("Root Cause")
    refresh = workbook.create_sheet("Refresh Control")
    guide = workbook.create_sheet("Workbook Guide")
    for cell, key in (
        ("B5", "observed_change_bps"),
        ("B6", "mix_contribution_bps"),
        ("B7", "within_segment_contribution_bps"),
        ("B8", "reconciliation_residual_bps"),
        ("B9", "primary_dimension"),
        ("B10", "primary_driver"),
        ("B12", "causal_status"),
    ):
        root[cell] = expected[key]
    refresh["B5"] = expected["evidence_id"]
    refresh["B6"] = expected["evidence_payload_sha256"]
    refresh["B7"] = expected["run_id"]
    refresh["B8"] = expected["reporting_period"]
    refresh["B11"] = expected["metric_registry_version"]
    refresh["B12"] = expected["data_quality_status"]
    guide["B9"] = "Synthetic default profile"
    workbook.save(path)
    validation = tmp_path / "outputs" / "validation" / "office_workbook_validation.json"
    validation.parent.mkdir(parents=True)
    validation.write_text(json.dumps({"status": "PASS"}), encoding="utf-8")
    _write_release_manifest(tmp_path, path, expected)

    result = _workbook_channel(tmp_path, expected)

    assert result["status"] == PASS
    assert not [item for item in result["checks"] if item["outcome"] != PASS]


def test_linkedin_missing_media_is_reported_as_incomplete_not_story_mismatch(
    tmp_path: Path,
) -> None:
    expected = _story()
    root = tmp_path / "outputs" / "linkedin"
    root.mkdir(parents=True)
    texts = {
        "research-summary.md": ("311.4150 4.4335 306.9815 Affiliate associational"),
        "technical-summary.md": "Synthetic evidence",
    }
    for name, content in texts.items():
        (root / name).write_text(content, encoding="utf-8")
    manifest = {
        "product": "nAIM Portfolio Intelligence Workbench",
        "tagline": "Name the movement. Own the evidence.",
        "evidence_id": expected["evidence_id"],
        "text_artifacts": {name: "READY" for name in texts},
    }
    (root / "package-manifest.json").write_text(json.dumps(manifest), encoding="utf-8")

    result = _linkedin_channel(tmp_path, expected)

    assert result["status"] == INCOMPLETE
    media_checks = [
        item
        for item in result["checks"]
        if item["check_id"]
        in {"linkedin.editable_carousel_present", "linkedin.pdf_carousel_present"}
    ]
    assert {item["outcome"] for item in media_checks} == {MISSING}
